#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
bwz-ebook.com 电子书信息采集工具

流程：
1. 界面填写开始时间、结束时间（按出版时间筛选书籍）
2. 时间范围按年拆分，逐年调用 https://www.bwz-ebook.com/getBookListApi 接口采集：
   每页固定返回 25 条书籍信息，通过返回的 count 计算总页数并翻页，
   每次调用接口后 sleep 0.5 秒
3. 各年数据按书籍ID去重后合并输出到同一 Excel，格式与《书籍信息采集模板.xlsx》一致

输出列：书籍ID(res_id)、书籍名称(res_name)、作者(author)、出版社(publishername)、
        出版时间(publish_date)、ISBN(isbn)、页数(留空)、书籍封面链接(thumbnail)、
        书籍链接(https://www.bwz-ebook.com/book/{res_id})、SSN号(留空)、读秀号(留空)
"""

import math
import os
import re
import threading
import time
import tkinter as tk
from datetime import datetime
from queue import Empty, Queue
from tkinter import scrolledtext, ttk

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill

# ==================== 配置区域 ====================

API_URL = "https://www.bwz-ebook.com/getBookListApi"
BOOK_URL_TPL = "https://www.bwz-ebook.com/book/{res_id}"
# 列表页 URL（Referer 中的时间参数需与界面一致）
BOOK_LIST_URL_TPL = (
    "https://www.bwz-ebook.com/bookList?isClear=true&resourcesType=book"
    "&startDate={start}&endDateVal={end}&startNumber=0&endNumber=99999999&bookType=1"
)

PAGE_SIZE = 25          # 接口每页固定返回 25 条
REQUEST_INTERVAL = 3  # 每次调用接口后 sleep 0.5 秒
API_TIMEOUT = 30        # 单次请求超时秒
API_RETRY = 3           # 网络异常重试次数
API_RETRY_SLEEP = 2     # 重试前等待秒数
FLUSH_ROWS = 500        # 每累计新增该条数，将 Excel 追加落盘一次（防意外丢失）

OUTPUT_DIR = "output"

# 输出列（与《书籍信息采集模板.xlsx》一致）
EXCEL_COLUMNS = ["书籍ID", "书籍名称", "作者", "出版社", "出版时间", "ISBN",
                 "页数", "书籍封面链接", "书籍链接", "SSN号", "读秀号"]

# ==================== 全局停止控制 ====================
_stop_event = threading.Event()


class TaskStoppedException(Exception):
    pass


def check_stop():
    if _stop_event.is_set():
        raise TaskStoppedException()


def interval_sleep(seconds: float):
    """可中断的 sleep：等待期间收到停止信号则立即抛出"""
    if _stop_event.wait(timeout=seconds):
        raise TaskStoppedException()


# ==================== 线程安全日志队列 ====================
LOG_QUEUE = Queue(maxsize=2000)


def log_print(text):
    LOG_QUEUE.put(text)


# ==================== 接口调用 ====================

def build_headers(start_date: str, end_date: str) -> dict:
    """请求头（Referer 中的 startDate/endDateVal 与界面一致）"""
    return {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36',
        'Content-Type': 'application/json',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br, zstd',
        'Connection': 'keep-alive',
        'Host': 'www.bwz-ebook.com',
        'Origin': 'https://www.bwz-ebook.com',
        'Referer': BOOK_LIST_URL_TPL.format(start=start_date, end=end_date),
    }


def build_req_data(start_date: str, end_date: str, page_index: int) -> dict:
    """请求体（startDate/endDateVal 与界面一致，pageIndex 为页码）"""
    return {
        "type": "ZHONGTU",
        "categorys": [],
        "institutionsName": [],
        "year": [],
        "pageIndex": page_index,
        "orderBy": "b.field25 desc , a.publish_date desc",
        "keyword": "",
        "isbn": "",
        "startDate": start_date,
        "endDateVal": end_date,
        "startNumber": "0",
        "endNumber": "99999999",
        "searchJudges": "",
        "searchTypes": "",
        "searchValues": "",
        "searchConditions": "",
        "isBuy": "0",
        "startTime": "",
        "endTime": "",
        "bookType": "1",
        "isClear": "true",
        "institutionIdDeleted": False,
        "startDateDeleted": False,
        "endDateValDeleted": False,
        "yearDeleted": False
    }


def fetch_page(start_date: str, end_date: str, page_index: int):
    """请求一页数据。成功返回 {"count": int, "items": list}，失败返回 None"""
    headers = build_headers(start_date, end_date)
    req_data = build_req_data(start_date, end_date, page_index)
    for attempt in range(1, API_RETRY + 1):
        check_stop()
        try:
            res = requests.post(API_URL, json=req_data, headers=headers, timeout=API_TIMEOUT)
            res.raise_for_status()
            j = res.json()
        except Exception as e:
            log_print(f"[!] 第 {page_index} 页请求失败（第{attempt}/{API_RETRY}次）: {e}")
            if attempt < API_RETRY:
                interval_sleep(API_RETRY_SLEEP)
            continue
        if j.get("code") != 0:
            log_print(f"[!] 接口返回异常: code={j.get('code')} message={j.get('message')}")
            return None
        return {"count": int(j.get("count") or 0), "items": j.get("data") or []}
    return None


def book_to_row(item: dict) -> list:
    """单条书籍数据 → 模板一行（页数、SSN号、读秀号接口无字段，留空）"""
    res_id = str(item.get("res_id") or "").strip()
    return [res_id,
            item.get("res_name") or "",
            item.get("author") or "",
            item.get("publishername") or "",
            item.get("publish_date") or "",
            item.get("isbn") or "",
            "",
            item.get("thumbnail") or "",
            BOOK_URL_TPL.format(res_id=res_id) if res_id else "",
            "",
            ""]


# ==================== Excel 输出 ====================

class ExcelWriter:
    def __init__(self, start_date: str, end_date: str, flush_every: int = FLUSH_ROWS):
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        self.flush_every = flush_every
        self.seen_ids = set()  # 已入库书籍ID，用于去重
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "书籍数据"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
            cell = self.ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        col_widths = [25, 40, 30, 25, 15, 20, 12, 50, 60, 15, 20]
        for i, width in enumerate(col_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            self.ws.column_dimensions[col_letter].width = width

        self.row_count = 0
        self.pending = 0  # 自上次落盘后新增的行数
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.path = os.path.join(
            OUTPUT_DIR, f"bwz_ebook_books_{start_date}_{end_date}_{stamp}.xlsx")

    def add_book(self, item: dict) -> bool:
        """书籍入库（按书籍ID去重）。返回 True 表示新增，False 表示重复/无效被跳过"""
        res_id = str(item.get("res_id") or "").strip()
        if not res_id or res_id in self.seen_ids:
            return False
        self.seen_ids.add(res_id)
        self.append(book_to_row(item))
        return True

    def append(self, row: list):
        self.ws.append(row)
        for cell in self.ws[self.ws.max_row]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        self.row_count += 1
        self.pending += 1
        # 每累计满 flush_every 条，追加落盘一次
        if self.pending >= self.flush_every:
            self.save()
            self.pending = 0
            log_print(f"[落盘] 已保存 {self.row_count} 条")

    def save(self):
        self.wb.save(self.path)


# ==================== 采集主流程 ====================

def build_year_segments(start_date: str, end_date: str) -> list:
    """把时间范围按年切分，返回 [(段开始, 段结束), ...] 列表。

    例：2023-05-01 ~ 2026-09-16 →
        [(2023-05-01, 2023-12-31), (2024-01-01, 2024-12-31),
         (2025-01-01, 2025-12-31), (2026-01-01, 2026-09-16)]
    """
    start_year = int(start_date[:4])
    end_year = int(end_date[:4])
    if start_year == end_year:
        return [(start_date, end_date)]
    segments = [(start_date, f"{start_year}-12-31")]
    for y in range(start_year + 1, end_year):
        segments.append((f"{y}-01-01", f"{y}-12-31"))
    segments.append((f"{end_year}-01-01", end_date))
    return segments


def collect_range(writer: ExcelWriter, start_date: str, end_date: str, label: str) -> bool:
    """采集一个时间段（通常为某一年）的书籍并写入 Excel。
    返回 True 表示该段处理完成（含无数据），False 表示中断。"""
    # 第 1 页：获取总数并计算总页数
    first = fetch_page(start_date, end_date, 1)
    if first is None:
        log_print(f"[!] {label} 首次请求失败，跳过该时间段")
        return False

    count = first["count"]
    total_pages = math.ceil(count / PAGE_SIZE) if count > 0 else 0
    log_print(f"[*] {label}: 符合条件的书籍共 {count} 本，{total_pages} 页")

    if count == 0:
        log_print(f"[*] {label}: 无数据，跳过")
        return True

    dup_total = 0  # 本段跳过的重复/无效书籍数
    for item in first["items"]:
        if not writer.add_book(item):
            dup_total += 1
    log_print(f"[采集] {label} 第 1/{total_pages} 页 本页{len(first['items'])}条 累计{writer.row_count}条")
    interval_sleep(REQUEST_INTERVAL)  # 每次调用后 sleep 0.5 秒

    for pg in range(2, total_pages + 1):
        check_stop()
        data = fetch_page(start_date, end_date, pg)
        if data is None:
            log_print(f"[!] {label} 第 {pg}/{total_pages} 页多次重试仍失败，该时间段中断（已采集数据已保存）")
            return False
        for item in data["items"]:
            if not writer.add_book(item):
                dup_total += 1
        if pg % 10 == 0 or pg == total_pages:
            log_print(f"[采集] {label} 第 {pg}/{total_pages} 页 本页{len(data['items'])}条 累计{writer.row_count}条")
        interval_sleep(REQUEST_INTERVAL)  # 每次调用后 sleep 0.5 秒
    if dup_total:
        log_print(f"[*] {label}: 本段共跳过重复/无效书籍 {dup_total} 条")
    return True


def run_collect(start_date: str, end_date: str):
    log_print("\n" + "=" * 65)
    log_print("  📚 开始书籍信息采集")
    log_print(f"  时间范围: {start_date} ~ {end_date}")
    log_print(f"  接口: {API_URL}  每页 {PAGE_SIZE} 条  间隔 {REQUEST_INTERVAL} 秒")
    log_print("=" * 65 + "\n")

    segments = build_year_segments(start_date, end_date)
    log_print(f"[*] 按年拆分为 {len(segments)} 段依次请求:")
    for s, e in segments:
        log_print(f"    {s} ~ {e}")

    writer = ExcelWriter(start_date, end_date)
    all_done = True
    try:
        for i, (s, e) in enumerate(segments, 1):
            check_stop()
            label = f"{s[:4]}年"
            log_print(f"\n===== [{i}/{len(segments)}] {label} ({s} ~ {e}) =====")
            if not collect_range(writer, s, e, label):
                all_done = False
                log_print(f"[!] {label} 采集不完整（已采集数据已保存，继续下一段）")
    finally:
        writer.save()

    log_print("\n" + "=" * 65)
    if all_done:
        log_print(f"[+] ✅ 采集完成！共 {writer.row_count} 条")
    else:
        log_print(f"[!] 采集存在中断，已保存 {writer.row_count} 条")
    log_print(f"[+] 输出文件: {os.path.abspath(writer.path)}")
    log_print("=" * 65 + "\n")


# ==================== GUI 界面 ====================

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("电子书信息采集工具（bwz-ebook）")
        self.root.geometry("800x640")
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        input_frame = ttk.Frame(root, padding=10)
        input_frame.pack(fill=tk.X)

        ttk.Label(input_frame, text="开始时间:").grid(row=0, column=0, sticky=tk.W, pady=3)
        self.start_var = tk.StringVar(value="2025-06-01")
        ttk.Entry(input_frame, textvariable=self.start_var, width=15).grid(row=0, column=1, padx=5)

        ttk.Label(input_frame, text="结束时间:").grid(row=0, column=2, sticky=tk.W, pady=3)
        self.end_var = tk.StringVar(value="2026-09-16")
        ttk.Entry(input_frame, textvariable=self.end_var, width=15).grid(row=0, column=3, padx=5)

        ttk.Label(input_frame, text="(格式: YYYY-MM-DD)").grid(row=0, column=4, sticky=tk.W, padx=5)

        btn_frame = ttk.Frame(input_frame)
        btn_frame.grid(row=1, column=0, columnspan=5, pady=8)
        self.start_btn = ttk.Button(btn_frame, text="开始执行", command=self.start_task)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        self.stop_btn = ttk.Button(btn_frame, text="结束执行", command=self.stop_task, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        ttk.Label(root, text="运行日志:", font=("Arial", 10, "bold")).pack(anchor=tk.W, padx=10, pady=(10, 0))
        self.log_text = scrolledtext.ScrolledText(root, height=26, state=tk.NORMAL, wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.running = False
        self.worker_thread = None
        self.consume_log_queue()

    def consume_log_queue(self):
        try:
            while True:
                msg = LOG_QUEUE.get_nowait()
                self.log_text.insert(tk.END, msg + "\n")
                self.log_text.see(tk.END)
        except Empty:
            pass
        self.root.after(50, self.consume_log_queue)

    def on_close(self):
        self.stop_task()
        self.root.destroy()

    def validate_dates(self):
        start = self.start_var.get().strip()
        end = self.end_var.get().strip()
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", start):
            log_print("[!] 错误：开始时间格式不正确，应为 YYYY-MM-DD")
            return None
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", end):
            log_print("[!] 错误：结束时间格式不正确，应为 YYYY-MM-DD")
            return None
        if start > end:
            log_print("[!] 错误：开始时间不能晚于结束时间")
            return None
        return start, end

    def start_task(self):
        if self.running:
            return
        dates = self.validate_dates()
        if not dates:
            return

        self.running = True
        _stop_event.clear()
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        self.worker_thread = threading.Thread(target=self.run_loop, args=dates, daemon=True)
        self.worker_thread.start()

    def stop_task(self):
        if not self.running:
            return
        self.running = False
        _stop_event.set()
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        log_print("\n[!] 已发送结束信号，正在中断当前操作...")

    def run_loop(self, start_date, end_date):
        try:
            run_collect(start_date, end_date)
        except TaskStoppedException:
            log_print("\n[!] 任务已被用户结束\n")
        except Exception as e:
            log_print(f"[!] 执行异常: {e}")
            import traceback
            log_print(traceback.format_exc())
        finally:
            self.running = False
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            log_print("\n[*] 本次执行结束\n")


def main():
    root = tk.Tk()
    app = App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
